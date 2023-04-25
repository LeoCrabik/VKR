import openpyxl as opxl
import random
from math import ceil


class ExcelReader:
    def __init__(self, path):
        self.path = path
        self.wb = opxl.load_workbook(self.path)
        self.sheet = self.wb.active
        self.student_row = 5
        self.names_dict = self.get_student_names()
        self.mark_names = {5: 'отлично', 4: 'хорошо', 3: 'удовлетворительно', 'зачет': 'зачтено'}
        self.subject_list = []
        self.full_subject_list = []
        self.marks_list = []
        self.current_block_name = 'Дисциплины'
        self.student_dict = {}
        self.course_projects = []
        self.birth_date = None
        self.subject_dict = self.get_subjects_dict()

    def reader(self):
        for item in self.names_dict.items():
            current_student_name, current_column = item
            mark_sum = 0
            mark_count = 0
            three_flag = False
            string_count = 0
            for row in range(1, self.sheet.max_row):

                current_subject = self.sheet.cell(row=row, column=3).value
                first_cell = self.sheet.cell(row=row, column=1).value
                if first_cell:
                    if first_cell in ['+', '-']:
                        current_mark = self.sheet.cell(row=row, column=current_column).value
                        try:
                            mark_sum += int(current_mark)
                            mark_count += 1
                            if current_mark == 3:
                                three_flag = True
                        except:
                            pass

                        if first_cell == '+':
                            if current_mark:
                                self.subject_list.append(current_subject)
                                string_count += ceil(len(current_subject) / 50) if current_subject else 1
                                # print(current_subject)
                                # print(string_count)
                                str_mark = self.mark_names[current_mark]
                                if self.sheet.cell(row=row, column=6).value:
                                    str_mark = 'зачтено ({})'.format(str_mark)
                                self.marks_list.append([current_subject, str_mark, 0 if string_count < 45 else 1])

                    elif '*' in first_cell:
                        self.course_projects.append(
                            [
                                self.sheet.cell(row=row, column=3).value,
                                self.sheet.cell(row=row, column=current_column).value.split(':')[0],
                                self.sheet.cell(row=row, column=current_column).value.split(':')[1],
                            ]
                        )
                    elif 'Год рождения' in first_cell:
                        self.birth_date = self.sheet.cell(row=row, column=current_column).value




                    else:
                        for block in ['Практики', 'Государственная итоговая аттестация', 'Факультативы']:
                            if block in first_cell:
                                self.current_block_name = block
                                self.marks_list.append([' ', ' ', ' ', 0 if string_count < 45 else 1])
                                self.marks_list.append([block, 'x', 0 if string_count < 45 else 1])
                                self.marks_list.append(['в том числе:', ' ', ' ', 0 if string_count < 45 else 1])


            # for subject in self.marks_list:
            #     print(' '.join(subject))
            # print('_' * 100)

            mean_mark = mark_sum / mark_count
            if not three_flag and mean_mark >= 4.75:
                mean_mark = ' с отличием'
            else:
                mean_mark = ''

            # for mark in self.marks_list:
            #     print(mark)

            self.student_dict[current_student_name] = (
                self.marks_list,
                self.course_projects,
                self.birth_date,
                mean_mark
            )
            self.marks_list = []
            self.current_block_name = 'Дисциплины'
            self.course_projects = []
            # current_column += 1
            # current_student_name = self.sheet.cell(row=5, column=current_column).value

        # self.subject_list = set(self.subject_list)
        # new_list = []
        # for row in range(1, self.sheet.max_row):
        #     if self.sheet.cell(row=row, column=1).value == '+':
        #         self.full_subject_list.append(self.sheet.cell(row=row, column=3).value)
        #         print(self.sheet.cell(row=row, column=3).value)
        # for subject in self.full_subject_list:
        #     if subject in self.subject_list:
        #         new_list.append(subject)
        # self.subject_list = new_list

        # self.subject_list = [self.sheet.cell(row=row, column=3).value for row in range(1, self.sheet.max_row) if (self.sheet.cell(row=row, column=1).value and '+' in self.sheet.cell(row=row, column=1).value and [self.sheet.cell(row=row, column=column).value for column in range(9, 9+len(self.student_dict.keys()))])]
        # print(self.subject_list)
        return self.student_dict

    def get_block_dict(self, i, block='some_string'):
        subject_list = []
        for row in range(i, self.sheet.max_row):
            if self.sheet.cell(row=row, column=1).value:
                if block in self.sheet.cell(row=row, column=1).value:
                    return [row, subject_list]
                elif '+' in self.sheet.cell(row=row, column=1).value and len([i for i in (self.sheet.cell(row=row, column=int(column)).value for column in self.names_dict.values()) if i is not None]) > 0:
                    if self.sheet.cell(row=row, column=3).value:
                        subject_list.append(self.sheet.cell(row=row, column=3).value)
        return [row, subject_list]

    def get_subjects_dict(self):
        subject_dict = {}
        row = 1
        block_list = ['Дисциплины', 'Практики', 'Государственная итоговая аттестация', 'Факультативы']
        for i in range(0, len(block_list)):
            try:
                row, subject_list = self.get_block_dict(row + 1, block_list[i+1])
            except:
                row, subject_list = self.get_block_dict(row+1)
            # print(row, subject_list)
            subject_dict[block_list[i]] = subject_list
        return subject_dict

    def get_student_names(self):
        names_dict = {}
        for column in range(9, self.sheet.max_column):
            current_student_name = self.sheet.cell(row=self.student_row, column=column).value
            # print(current_student_name)
            if not current_student_name:
                return names_dict
            else:
                names_dict[current_student_name] = column


if __name__ == '__main__':
    excel = ExcelReader(path='C:/Users/leocr/PycharmProjects/vkr/vkr/excel/book1.xlsx')
    # print(excel.reader()['Боков Даниил Александрович'][1])
    for subject in excel.subject_dict.items():
        print(subject)
    print(excel.get_student_names())
